# repair_excel_table.py
import os, shutil, time
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

EXCEL_PATH = "data.xlsx"
BACKUP_DIR = "backups"

def backup_file(path):
    os.makedirs(BACKUP_DIR, exist_ok=True)
    stamp = time.strftime("%Y%m%d_%H%M%S")
    name = os.path.basename(path)
    bname = f"{name}.bak.{stamp}"
    dest = os.path.join(BACKUP_DIR, bname)
    shutil.copy2(path, dest)
    return dest

def sanitize_workbook(path):
    if not os.path.exists(path):
        print("No data.xlsx found; nothing to repair.")
        return

    print("Backing up original file...")
    backup = backup_file(path)
    print("Backup written to:", backup)

    wb = load_workbook(path)
    repaired = False

    for ws in wb.worksheets:
        print(f"Inspecting worksheet: {ws.title}")
        # Filter ws._tables to only objects that look like real Table instances
        real_tables = [t for t in ws._tables if hasattr(t, "ref") and hasattr(t, "displayName")]
        bogus_count = len(ws._tables) - len(real_tables)
        if bogus_count > 0:
            print(f"  - Removing {bogus_count} non-Table entries from ws._tables.")
            ws._tables = real_tables
            repaired = True
        # If no real table exists, create one assuming header row at row 1 and 3 columns (A:C)
        if len(real_tables) == 0:
            max_row = ws.max_row
            if max_row == 0:
                print("  - sheet empty; skipping table creation.")
                continue
            # If there are fewer than 3 columns, still create table using available columns
            max_col = min(3, ws.max_column)
            col_letter_end = ['A','B','C'][max_col-1] if max_col >= 1 else 'A'
            ref = f"A1:{col_letter_end}{max_row}"
            # ensure table name is unique
            existing_names = {getattr(t, "displayName", "") for t in ws._tables}
            base = "DataTable"
            idx = 1
            name = base
            while name in existing_names:
                idx += 1
                name = f"{base}{idx}"
            table = Table(displayName=name, ref=ref)
            style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
            table.tableStyleInfo = style
            ws.add_table(table)
            print(f"  - Added table {name} with ref {ref}")
            repaired = True
        else:
            # update every real table's ref to ensure it covers all rows A1:C<max_row>
            max_row = ws.max_row
            if max_row >= 1:
                new_ref = f"A1:C{max_row}"
                for t in ws._tables:
                    try:
                        if getattr(t, "ref", None) != new_ref:
                            print(f"  - Updating table {t.displayName} ref -> {new_ref}")
                            t.ref = new_ref
                            repaired = True
                    except Exception as ex:
                        print(f"  - Could not update table {getattr(t,'displayName', '<unknown>')}: {ex}")

    if repaired:
        wb.save(path)
        print("Repaired workbook saved to", path)
    else:
        print("No repairs necessary; workbook left unchanged.")

if __name__ == "__main__":
    sanitize_workbook(EXCEL_PATH)
